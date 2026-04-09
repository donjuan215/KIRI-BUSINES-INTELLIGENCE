from flask import Flask, render_template, request, redirect, session, jsonify
import os
import json
from datetime import datetime
from dotenv import load_dotenv
load_dotenv() 
TENANT_ID = "store_001"
TENANT_PATH = os.path.join("tenants", TENANT_ID)

app = Flask(__name__)
app.secret_key = "kyri_secret_2026"


def load_json(filename):
    path = os.path.join(TENANT_PATH, filename)
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(filename, data):
    path = os.path.join(TENANT_PATH, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def is_logged_in():
    return session.get("logged_in", False)


def get_dashboard_data():
    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})
    meta_diaria = config.get("daily_goal", 0)

    total_ingresos = 0
    total_costo = 0
    total_ganancia = 0
    ventas_lista = []
    stock_critico = []
    stock_ok = []

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        stock = p.get("stock", 0)
        ingreso = vendidos * precio
        ganancia = vendidos * (precio - costo)
        total_ingresos = total_ingresos + ingreso
        total_costo = total_costo + (vendidos * costo)
        total_ganancia = total_ganancia + ganancia

        if vendidos > 0:
            dias = round(stock / vendidos, 1)
        else:
            dias = 999

        if vendidos > 0:
            ventas_lista.append({
                "id": pid,
                "nombre": p.get("name", pid),
                "vendidos": vendidos,
                "ingreso": ingreso,
                "precio": precio,
                "dias": dias,
                "stock": stock,
                "pct_bar": 0
            })

        item = {"nombre": p.get("name", pid), "stock": stock, "dias": dias}
        if stock <= 5:
            item["nivel"] = "red"
            stock_critico.append(item)
        elif stock <= 15:
            item["nivel"] = "yellow"
            stock_critico.append(item)
        else:
            item["nivel"] = "green"
            stock_ok.append(item)

    ventas_lista.sort(key=lambda x: x["ingreso"], reverse=True)

    if total_ingresos > 0:
        margen = round(total_ganancia / total_ingresos * 100, 1)
    else:
        margen = 0

    if meta_diaria > 0:
        pct_meta = round(total_ingresos / meta_diaria * 100, 1)
    else:
        pct_meta = 0

    if len(ventas_lista) > 0:
        max_ingreso = ventas_lista[0]["ingreso"]
    else:
        max_ingreso = 1

    for p in ventas_lista:
        if max_ingreso > 0:
            p["pct_bar"] = round(p["ingreso"] / max_ingreso * 100)

    return {
        "business_name": config.get("name", TENANT_ID),
        "total_ingresos": total_ingresos,
        "total_costo": total_costo,
        "total_ganancia": total_ganancia,
        "margen": margen,
        "pct_meta": pct_meta,
        "meta_diaria": meta_diaria,
        "ventas_lista": ventas_lista,
        "stock_critico": stock_critico,
        "stock_ok": stock_ok,
        "insights": state.get("insights", []),
        "alerts": state.get("alerts", []),
        "fecha": datetime.now().strftime("%d/%m/%Y - %I:%M %p"),
        "n_criticos": len(stock_critico)
    }


def guardar_historial_diario():
    hoy = datetime.now().strftime("%Y-%m-%d")
    historial_dir = os.path.join(TENANT_PATH, "historial")
    if not os.path.exists(historial_dir):
        os.makedirs(historial_dir)
    state = load_json("state.json")
    config = load_json("config.json")
    snapshot = {
        "fecha": hoy,
        "hora_guardado": datetime.now().strftime("%I:%M %p"),
        "business_name": config.get("name", TENANT_ID),
        "products": state.get("products", {}),
        "sales": state.get("sales", 0),
        "historial_ventas": state.get("historial_ventas", []),
        "daily_goal": config.get("daily_goal", 0)
    }
    path = os.path.join(historial_dir, hoy + ".json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)
    print("Historial guardado: " + path)


def generar_alertas_y_consejos(state, config):
    productos = state.get("products", {})
    meta_diaria = config.get("daily_goal", 0)

    total_ingresos = 0
    total_ganancia = 0
    alertas = []
    consejos = []
    criticos = []
    bajos = []

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        stock = p.get("stock", 0)
        ingreso = vendidos * precio
        ganancia = vendidos * (precio - costo)
        total_ingresos = total_ingresos + ingreso
        total_ganancia = total_ganancia + ganancia

        if stock <= 5:
            criticos.append(p.get("name", pid))
        elif stock <= 15:
            bajos.append(p.get("name", pid))

    if total_ingresos > 0:
        margen = round(total_ganancia / total_ingresos * 100, 1)
    else:
        margen = 0

    if meta_diaria > 0:
        pct_meta = round(total_ingresos / meta_diaria * 100, 1)
    else:
        pct_meta = 0

    if len(criticos) > 0:
        alertas.append({
            "tipo": "Stock Crítico",
            "titulo": "Productos agotándose",
            "descripcion": "Tienes " + str(len(criticos)) + " producto(s) con stock crítico: " + ", ".join(criticos) + ". Debes reabastecer pronto.",
            "badge": str(len(criticos)) + " producto(s)",
            "icono": "🔴",
            "color": "red"
        })

    if len(bajos) > 0:
        alertas.append({
            "tipo": "Stock Bajo",
            "titulo": "Stock por agotarse",
            "descripcion": "Estos productos tienen stock bajo: " + ", ".join(bajos) + ". Considera reabastecer esta semana.",
            "badge": str(len(bajos)) + " producto(s)",
            "icono": "🟡",
            "color": "orange"
        })

    if meta_diaria > 0 and pct_meta >= 100:
        alertas.append({
            "tipo": "Meta Alcanzada",
            "titulo": "Meta del día superada",
            "descripcion": "Alcanzaste el " + str(pct_meta) + "% de tu meta diaria. Excelente rendimiento hoy.",
            "badge": str(pct_meta) + "% completado",
            "icono": "🎯",
            "color": "green"
        })
    elif meta_diaria > 0 and pct_meta >= 50:
        alertas.append({
            "tipo": "Meta en Progreso",
            "titulo": "Vas por buen camino",
            "descripcion": "Llevas el " + str(pct_meta) + "% de tu meta. Sigue así para cerrar el día en verde.",
            "badge": str(pct_meta) + "% completado",
            "icono": "📈",
            "color": "blue"
        })

    if margen < 20 and total_ingresos > 0:
        alertas.append({
            "tipo": "Margen Bajo",
            "titulo": "Margen de ganancia bajo",
            "descripcion": "Tu margen actual es " + str(margen) + "%. Revisa los costos o ajusta precios para mejorar la rentabilidad.",
            "badge": str(margen) + "% margen",
            "icono": "⚠",
            "color": "orange"
        })

    if len(criticos) > 0:
        consejos.append({
            "tipo": "inventario",
            "titulo": "Reabastece " + criticos[0] + " hoy",
            "texto": "Con el ritmo actual de ventas, " + criticos[0] + " se agota muy pronto. Haz el pedido antes de que afecte las ventas.",
            "icono": "📦", "color": "red", "accion": "Ver inventario", "link": "/inventory"
        })

    if total_ingresos > 0 and pct_meta < 60 and meta_diaria > 0:
        faltante = meta_diaria - total_ingresos
        consejos.append({
            "tipo": "ventas",
            "titulo": "Activa una promocion para llegar a la meta",
            "texto": "Te faltan $" + "{:,.0f}".format(faltante) + " COP para tu meta. Considera un combo o descuento en las proximas horas para impulsar las ventas.",
            "icono": "🎯", "color": "orange", "accion": "Ver ventas", "link": "/ventas"
        })

    if margen >= 40 and total_ingresos > 0:
        consejos.append({
            "tipo": "rentabilidad",
            "titulo": "Excelente rentabilidad hoy",
            "texto": "Tu margen del " + str(margen) + "% esta por encima del promedio del sector. Los productos con mayor margen son tu motor de ganancia — priorizalos.",
            "icono": "💰", "color": "green", "accion": None, "link": ""
        })

    if len(bajos) > 0:
        consejos.append({
            "tipo": "inventario",
            "titulo": "Planifica reabastecimiento",
            "texto": str(len(bajos)) + " productos tienen stock bajo: " + ", ".join(bajos) + ". Con base en tu ritmo de ventas, haz el pedido en los proximos 2 dias.",
            "icono": "🔄", "color": "blue", "accion": "Ver inventario", "link": "/inventory"
        })

    if total_ingresos == 0:
        consejos.append({
            "tipo": "inicio",
            "titulo": "Empieza a registrar ventas",
            "texto": "Registra tus ventas del dia desde el POS para que KYRI pueda analizar tu negocio y darte consejos personalizados en tiempo real.",
            "icono": "🚀", "color": "blue", "accion": "Ir a ventas", "link": "/ventas"
        })

    return alertas, consejos



# ---------- LOGIN ----------

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        config = load_json("config.json")
        usuario = request.form.get("usuario", "").strip()
        clave = request.form.get("clave", "").strip()
        if usuario == config.get("usuario", "admin") and clave == config.get("clave", "1234"):
            session["logged_in"] = True
            session["business_name"] = config.get("name", TENANT_ID)
            return redirect("/dashboard")
        else:
            error = "Usuario o contrasena incorrectos"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ---------- DASHBOARD ----------
import smtplib
import random
import string
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def generar_credenciales():
    usuario = "kyri_" + ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
    clave = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
    return usuario, clave

def crear_tenant(nombre, negocio, tipo, email, ciudad, wsp):
    tenant_id = "store_" + ''.join(random.choices(string.digits, k=6))
    tenant_path = os.path.join("tenants", tenant_id)
    historial_path = os.path.join(tenant_path, "historial")
    os.makedirs(tenant_path, exist_ok=True)
    os.makedirs(historial_path, exist_ok=True)
    usuario, clave = generar_credenciales()
    config = {
        "name": negocio, "tipo": tipo, "owner": nombre,
        "email": email, "ciudad": ciudad, "whatsapp": wsp,
        "usuario": usuario, "clave": clave,
        "daily_goal": 0, "moneda": "COP"
    }
    with open(os.path.join(tenant_path, "config.json"), "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    state = {"products": {}, "sales": 0, "historial_ventas": [], "insights": []}
    with open(os.path.join(tenant_path, "state.json"), "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)
    return tenant_id, usuario, clave

def enviar_correo_bienvenida(nombre, email, usuario, clave, negocio):
    try:
        gmail_user = os.environ.get("GMAIL_USER")
        gmail_pass = os.environ.get("GMAIL_PASS")
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "🟢 Tu acceso a KYRI está listo"
        msg["From"] = f"KYRI Business Intelligence <{gmail_user}>"
        msg["To"] = email
        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="UTF-8"></head>
        <body style="margin:0;padding:0;background:#f5f7fa;font-family:Arial,sans-serif;">
          <div style="max-width:560px;margin:40px auto;background:#ffffff;border:1px solid #e4e8ef;border-radius:16px;overflow:hidden;">
            <div style="background:#0f1923;padding:28px 36px;">
              <span style="font-size:26px;font-weight:800;color:#ffffff;letter-spacing:-1px;">KYR<span style="color:#00e5a0;">I</span></span>
              <span style="font-size:9px;color:#7a8a99;letter-spacing:3px;text-transform:uppercase;margin-left:10px;">Business Intelligence</span>
            </div>
            <div style="height:3px;background:linear-gradient(90deg,#00966a,#0052cc);"></div>
            <div style="padding:36px;">
              <div style="font-size:11px;color:#00966a;text-transform:uppercase;letter-spacing:2px;font-weight:600;margin-bottom:10px;">Acceso activado</div>
              <div style="font-size:24px;font-weight:800;color:#0f1923;margin-bottom:8px;">¡Bienvenido a KYRI, {nombre}!</div>
              <div style="font-size:14px;color:#4a5568;line-height:1.7;margin-bottom:28px;">Tu negocio <strong>{negocio}</strong> ya está registrado. Aquí están tus credenciales:</div>
              <div style="background:#f0f2f5;border:1px solid #e4e8ef;border-radius:12px;padding:24px;margin-bottom:28px;">
                <div style="margin-bottom:16px;">
                  <div style="font-size:11px;color:#9aa5b4;text-transform:uppercase;letter-spacing:1px;margin-bottom:5px;">Usuario</div>
                  <div style="font-size:18px;font-weight:800;color:#0f1923;background:#ffffff;border:1px solid #e4e8ef;border-radius:8px;padding:10px 16px;">{usuario}</div>
                </div>
                <div>
                  <div style="font-size:11px;color:#9aa5b4;text-transform:uppercase;letter-spacing:1px;margin-bottom:5px;">Contraseña</div>
                  <div style="font-size:18px;font-weight:800;color:#0f1923;background:#ffffff;border:1px solid #e4e8ef;border-radius:8px;padding:10px 16px;letter-spacing:2px;">{clave}</div>
                </div>
              </div>
              <a href="http://localhost:5000/login" style="display:block;text-align:center;padding:16px;background:#00966a;color:#ffffff;border-radius:10px;font-size:15px;font-weight:700;text-decoration:none;margin-bottom:24px;">Entrar a KYRI →</a>
              <div style="border-top:1px solid #e4e8ef;padding-top:22px;">
                <div style="font-size:13px;font-weight:700;color:#0f1923;margin-bottom:12px;">¿Por dónde empezar?</div>
                <div style="font-size:13px;color:#4a5568;margin-bottom:8px;">📦 <strong>Inventario</strong> — Agrega tus productos con precio y costo</div>
                <div style="font-size:13px;color:#4a5568;margin-bottom:8px;">🎯 <strong>Configuración</strong> — Define tu meta diaria de ventas</div>
                <div style="font-size:13px;color:#4a5568;margin-bottom:8px;">📈 <strong>Ventas</strong> — Registra cada venta con el POS</div>
                <div style="font-size:13px;color:#4a5568;">🤖 <strong>KYRI Chat</strong> — Pregúntale cualquier cosa sobre tu negocio</div>
              </div>
            </div>
            <div style="background:#f5f7fa;border-top:1px solid #e4e8ef;padding:20px 36px;text-align:center;">
              <div style="font-size:11px;color:#9aa5b4;">KYRI Business Intelligence · Colombia</div>
            </div>
          </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(html, "html"))
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(gmail_user, gmail_pass)
        server.sendmail(gmail_user, email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False

@app.route("/registro", methods=["POST"])
def registro():
    body = request.get_json(force=True, silent=True)
    if not body:
        return jsonify({"ok": False, "error": "Sin datos"})
    nombre    = body.get("nombre", "").strip()
    negocio   = body.get("negocio", "").strip()
    tipo      = body.get("tipo", "").strip()
    email     = body.get("email", "").strip()
    ciudad    = body.get("ciudad", "").strip()
    productos = body.get("productos", "").strip()
    actual    = body.get("actual", "").strip()
    wsp       = body.get("wsp", "").strip()
    if not all([nombre, negocio, tipo, email, ciudad, wsp]):
        return jsonify({"ok": False, "error": "Faltan campos"})
    try:
        tenant_id, usuario, clave = crear_tenant(nombre, negocio, tipo, email, ciudad, wsp)
        correo_enviado = enviar_correo_bienvenida(nombre, email, usuario, clave, negocio)
        return jsonify({"ok": True, "correo": correo_enviado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
@app.route("/")
def home():
    if is_logged_in():
        return redirect("/dashboard")
    return render_template("landing.html")


@app.route("/dashboard")
def dashboard():
    if not is_logged_in():
        return redirect("/login")
    data = get_dashboard_data()
    return render_template("dashboard.html", **data)


# ---------- INVENTARIO ----------

@app.route("/inventory")
def inventory():
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})
    lista = []
    for pid in productos:
        p = productos[pid]
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        if precio > 0:
            margen = round((precio - costo) / precio * 100, 1)
        else:
            margen = 0
        lista.append({
            "id": pid,
            "nombre": p.get("name", pid),
            "precio": precio,
            "costo": costo,
            "stock": p.get("stock", 0),
            "vendidos": p.get("sold", 0),
            "margen": margen
        })
    return render_template("inventory.html",
        productos=lista,
        business_name=config.get("name", TENANT_ID),
        fecha=datetime.now().strftime("%d/%m/%Y - %I:%M %p")
    )


@app.route("/inventory/add", methods=["POST"])
def add_product():
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    productos = state.get("products", {})
    nuevo_id = str(len(productos) + 1)
    while nuevo_id in productos:
        nuevo_id = str(int(nuevo_id) + 1)
    productos[nuevo_id] = {
        "name": request.form["nombre"].strip(),
        "price": float(request.form["precio"]),
        "cost": float(request.form["costo"]),
        "stock": int(request.form["stock"]),
        "sold": 0
    }
    state["products"] = productos
    save_json("state.json", state)
    return redirect("/inventory")


@app.route("/inventory/edit/<pid>", methods=["POST"])
def edit_product(pid):
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    productos = state.get("products", {})
    if pid in productos:
        productos[pid]["name"] = request.form["nombre"].strip()
        productos[pid]["price"] = float(request.form["precio"])
        productos[pid]["cost"] = float(request.form["costo"])
        productos[pid]["stock"] = int(request.form["stock"])
    state["products"] = productos
    save_json("state.json", state)
    return redirect("/inventory")


@app.route("/inventory/delete/<pid>", methods=["POST"])
def delete_product(pid):
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    productos = state.get("products", {})
    if pid in productos:
        del productos[pid]
    state["products"] = productos
    save_json("state.json", state)
    return redirect("/inventory")


# ---------- VENTAS ----------

@app.route("/ventas")
def ventas():
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})
    lista = []
    for pid in productos:
        p = productos[pid]
        if p.get("stock", 0) > 0:
            lista.append({
                "id": pid,
                "nombre": p.get("name", pid),
                "precio": p.get("price", 0),
                "stock": p.get("stock", 0)
            })
    historial_raw = state.get("historial_ventas", [])
    ultimos = historial_raw[-20:]
    ultimos.reverse()
    return render_template("ventas.html",
        productos=lista,
        historial=ultimos,
        business_name=config.get("name", TENANT_ID),
        fecha=datetime.now().strftime("%d/%m/%Y - %I:%M %p"),
        total_dia=state.get("sales", 0)
    )


@app.route("/ventas/registrar", methods=["POST"])
def registrar_venta():
    if not is_logged_in():
        return jsonify({"ok": False, "error": "No autorizado"})

    state = load_json("state.json")
    productos = state.get("products", {})

    body = request.get_json(force=True, silent=True)
    if body is None:
        return jsonify({"ok": False, "error": "No se recibieron datos"})

    carrito = body.get("carrito", [])
    if len(carrito) == 0:
        return jsonify({"ok": False, "error": "Carrito vacio"})

    total_venta = 0
    items_vendidos = []
    total_cantidad = 0

    for item in carrito:
        pid = str(item["id"])
        cantidad = int(item["cantidad"])
        if pid not in productos:
            continue
        p = productos[pid]
        if p["stock"] < cantidad:
            return jsonify({"ok": False, "error": "Stock insuficiente para " + p["name"]})
        subtotal = cantidad * p["price"]
        total_venta = total_venta + subtotal
        total_cantidad = total_cantidad + cantidad
        productos[pid]["stock"] = productos[pid]["stock"] - cantidad
        productos[pid]["sold"] = productos[pid].get("sold", 0) + cantidad
        items_vendidos.append({
            "nombre": p["name"],
            "cantidad": cantidad,
            "precio": p["price"],
            "subtotal": subtotal
        })

    state["products"] = productos
    state["sales"] = state.get("sales", 0) + total_cantidad

    historial = state.get("historial_ventas", [])
    historial.append({
        "hora": datetime.now().strftime("%I:%M %p"),
        "items": items_vendidos,
        "total": total_venta
    })
    state["historial_ventas"] = historial
    save_json("state.json", state)
    config = load_json("config.json")
    numero_ticket = len(historial)
    return jsonify({
        "ok": True,
        "total": total_venta,
        "mensaje": "Venta registrada",
        "items": items_vendidos,
        "business_name": config.get("name", TENANT_ID),
        "fecha": datetime.now().strftime("%d/%m/%Y"),
        "hora": datetime.now().strftime("%I:%M %p"),
        "numero_ticket": numero_ticket
    })


# ---------- TIENDA ----------

@app.route("/tienda")
def tienda():
    if not is_logged_in():
        return redirect("/login")
    config = load_json("config.json")
    guardado = request.args.get("guardado", False)
    return render_template("tienda.html",
        config=config,
        business_name=config.get("name", TENANT_ID),
        fecha=datetime.now().strftime("%d/%m/%Y - %I:%M %p"),
        guardado=guardado
    )


@app.route("/tienda/guardar", methods=["POST"])
def guardar_tienda():
    if not is_logged_in():
        return redirect("/login")
    config = load_json("config.json")
    seccion = request.form.get("seccion", "")

    if seccion == "negocio":
        config["name"] = request.form["name"].strip()
        config["type"] = request.form.get("type", "").strip()
        config["daily_goal"] = float(request.form.get("daily_goal", 0))
        config["whatsapp"] = request.form.get("whatsapp", "").strip()
        session["business_name"] = config["name"]

    elif seccion == "acceso":
        config["usuario"] = request.form["usuario"].strip()
        nueva_clave = request.form.get("clave", "").strip()
        if nueva_clave:
            config["clave"] = nueva_clave

    save_json("config.json", config)
    return redirect("/tienda?guardado=1")


@app.route("/tienda/reset", methods=["POST"])
def reset_ventas():
    if not is_logged_in():
        return redirect("/login")
    guardar_historial_diario()  # ✅ guarda antes de borrar
    state = load_json("state.json")
    productos = state.get("products", {})
    for pid in productos:
        productos[pid]["sold"] = 0
    state["products"] = productos
    state["sales"] = 0
    state["historial_ventas"] = []
    save_json("state.json", state)
    return redirect("/tienda?guardado=1")


# ---------- ALERTAS ----------

@app.route("/alertas")
def alertas():
    if not is_logged_in():
        return redirect("/login")
    state = load_json("state.json")
    config = load_json("config.json")
    alertas_activas, consejos = generar_alertas_y_consejos(state, config)
    return render_template("alertas.html",
        alertas_activas=alertas_activas,
        consejos=consejos,
        business_name=config.get("name", TENANT_ID),
        fecha=datetime.now().strftime("%d/%m/%Y - %I:%M %p")
    )


# ---------- KYRI CHAT ----------

@app.route("/kyri_chat", methods=["POST"])
def kyri_chat():
    if not is_logged_in():
        return jsonify({"ok": False, "error": "No autorizado"})

    body = request.get_json(force=True, silent=True)
    if not body:
        return jsonify({"ok": False, "error": "Sin datos"})

    mensaje = body.get("mensaje", "").strip()
    if not mensaje:
        return jsonify({"ok": False, "error": "Mensaje vacio"})

    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})

    total_ingresos = 0
    total_ganancia = 0
    resumen_productos = []

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        stock = p.get("stock", 0)
        ingreso = vendidos * precio
        ganancia = vendidos * (precio - costo)
        total_ingresos = total_ingresos + ingreso
        total_ganancia = total_ganancia + ganancia
        resumen_productos.append(p.get("name", pid) + ": " + str(vendidos) + " vendidos, $" + "{:,.0f}".format(ingreso) + " ingresos, stock: " + str(stock) + " uds")

    if total_ingresos > 0:
        margen = round(total_ganancia / total_ingresos * 100, 1)
    else:
        margen = 0

    meta_diaria = config.get("daily_goal", 0)
    if meta_diaria > 0:
        pct_meta = round(total_ingresos / meta_diaria * 100, 1)
    else:
        pct_meta = 0

    contexto = (
        "Eres KYRI, asistente de inteligencia de negocios para comercios colombianos. "
        "Eres directo, amigable y das consejos practicos. Respondes en español. "
        "Nunca inventas datos. Maximo 3 oraciones cortas.\n\n"
        "DATOS DEL NEGOCIO HOY:\n"
        "Negocio: " + config.get("name", TENANT_ID) + "\n"
        "Ingresos: $" + "{:,.0f}".format(total_ingresos) + " COP\n"
        "Ganancia: $" + "{:,.0f}".format(total_ganancia) + " COP\n"
        "Margen: " + str(margen) + "%\n"
        "Meta: $" + "{:,.0f}".format(meta_diaria) + " COP (" + str(pct_meta) + "% completado)\n"
        "Transacciones: " + str(len(state.get("historial_ventas", []))) + "\n\n"
        "PRODUCTOS:\n" + "\n".join(resumen_productos)
    )

    try:
        from groq import Groq
        client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
        respuesta = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            max_tokens=300,
            messages=[
                {"role": "system", "content": contexto},
                {"role": "user", "content": mensaje}
            ]
        )
        texto = respuesta.choices[0].message.content
        return jsonify({"ok": True, "respuesta": texto})
    except Exception as e:
        return jsonify({"ok": False, "error": "Error IA: " + str(e)})


# ---------- REPORTE ----------

@app.route("/send_report", methods=["POST"])
def send_report():
    if not is_logged_in():
        return jsonify({"message": "No autorizado"})
    try:
        import sys
        sys.path.insert(0, ".")
        from core.insights.daily_report import generate_daily_report
        from core.notifications.whatsapp_sender import send_whatsapp
        state = load_json("state.json")
        config = load_json("config.json")
        state["business_name"] = config.get("name", TENANT_ID)
        state["daily_goal"] = config.get("daily_goal", 0)
        reporte = generate_daily_report(state, TENANT_ID)
        send_whatsapp(reporte)
        return jsonify({"message": "Reporte enviado por WhatsApp"})
    except Exception as e:
        return jsonify({"message": "Error: " + str(e)})


# ---------- EXPORTAR EXCEL ----------

@app.route("/exportar/excel")
def exportar_excel():
    if not is_logged_in():
        return redirect("/login")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io

    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})
    historial = state.get("historial_ventas", [])
    hoy = datetime.now().strftime("%d/%m/%Y")
    business = config.get("name", TENANT_ID)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Resumen del Día"

    verde = "00966a"
    gris = "f0f2f5"
    oscuro = "0f1923"

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    ws1["A1"] = "KYRI — Reporte Diario"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor=oscuro)
    ws1.merge_cells("A1:B1")

    ws1["A2"] = business + " — " + hoy
    ws1["A2"].font = Font(size=11, color="FFFFFF")
    ws1["A2"].fill = PatternFill("solid", fgColor=verde)
    ws1.merge_cells("A2:B2")

    total_ingresos = 0
    total_ganancia = 0
    total_costo = 0

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        total_ingresos += vendidos * precio
        total_costo += vendidos * costo
        total_ganancia += vendidos * (precio - costo)

    margen = round(total_ganancia / total_ingresos * 100, 1) if total_ingresos > 0 else 0
    meta = config.get("daily_goal", 0)
    pct_meta = round(total_ingresos / meta * 100, 1) if meta > 0 else 0

    kpis = [
        ("Ingresos Totales", "$" + "{:,.0f}".format(total_ingresos) + " COP"),
        ("Ganancia Neta", "$" + "{:,.0f}".format(total_ganancia) + " COP"),
        ("Costo de Ventas", "$" + "{:,.0f}".format(total_costo) + " COP"),
        ("Margen de Ganancia", str(margen) + "%"),
        ("Meta del Día", "$" + "{:,.0f}".format(meta) + " COP"),
        ("Cumplimiento Meta", str(pct_meta) + "%"),
        ("Total Transacciones", str(len(historial))),
    ]

    ws1["A4"] = "INDICADORES"
    ws1["A4"].font = Font(bold=True, color="FFFFFF")
    ws1["A4"].fill = PatternFill("solid", fgColor=verde)
    ws1["B4"] = "VALOR"
    ws1["B4"].font = Font(bold=True, color="FFFFFF")
    ws1["B4"].fill = PatternFill("solid", fgColor=verde)

    for i, (label, valor) in enumerate(kpis):
        row = 5 + i
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=valor)
        if i % 2 == 0:
            ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor=gris)
            ws1.cell(row=row, column=2).fill = PatternFill("solid", fgColor=gris)

    ws2 = wb.create_sheet("Productos")
    for w, col in zip([5,25,15,15,15,12,12], "ABCDEFG"):
        ws2.column_dimensions[col].width = w

    headers = ["#", "Producto", "Precio", "Costo", "Margen", "Vendidos", "Stock"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=oscuro)

    lista = []
    for pid in productos:
        p = productos[pid]
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        vendidos = p.get("sold", 0)
        ingreso = vendidos * precio
        mg = round((precio - costo) / precio * 100, 1) if precio > 0 else 0
        lista.append((p.get("name", pid), precio, costo, mg, vendidos, p.get("stock", 0), ingreso))

    lista.sort(key=lambda x: x[6], reverse=True)

    for i, (nombre, precio, costo, mg, vendidos, stock, ingreso) in enumerate(lista):
        row = 2 + i
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=nombre)
        ws2.cell(row=row, column=3, value="$" + "{:,.0f}".format(precio))
        ws2.cell(row=row, column=4, value="$" + "{:,.0f}".format(costo))
        ws2.cell(row=row, column=5, value=str(mg) + "%")
        ws2.cell(row=row, column=6, value=vendidos)
        ws2.cell(row=row, column=7, value=stock)
        if i % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor=gris)

    ws3 = wb.create_sheet("Historial Ventas")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 18

    for col, h in enumerate(["Hora", "Productos", "Total"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=oscuro)

    for i, venta in enumerate(reversed(historial)):
        row = 2 + i
        items_txt = ", ".join([str(it.get("cantidad", 1)) + "x " + it.get("nombre", "") for it in venta.get("items", [])])
        ws3.cell(row=row, column=1, value=venta.get("hora", ""))
        ws3.cell(row=row, column=2, value=items_txt)
        ws3.cell(row=row, column=3, value="$" + "{:,.0f}".format(venta.get("total", 0)))
        if i % 2 == 0:
            for col in range(1, 4):
                ws3.cell(row=row, column=col).fill = PatternFill("solid", fgColor=gris)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = "KYRI_Reporte_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nombre_archivo)


# ---------- EXPORTAR PDF ----------

@app.route("/exportar/pdf")
def exportar_pdf():
    if not is_logged_in():
        return redirect("/login")
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from flask import send_file
    import io

    state = load_json("state.json")
    config = load_json("config.json")
    productos = state.get("products", {})
    historial = state.get("historial_ventas", [])
    hoy = datetime.now().strftime("%d/%m/%Y")
    business = config.get("name", TENANT_ID)

    total_ingresos = 0
    total_ganancia = 0
    total_costo = 0

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        total_ingresos += vendidos * precio
        total_costo += vendidos * costo
        total_ganancia += vendidos * (precio - costo)

    margen = round(total_ganancia / total_ingresos * 100, 1) if total_ingresos > 0 else 0
    meta = config.get("daily_goal", 0)
    pct_meta = round(total_ingresos / meta * 100, 1) if meta > 0 else 0

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=2*cm, rightMargin=2*cm)

    verde = colors.HexColor("#00966a")
    oscuro = colors.HexColor("#0f1923")
    gris = colors.HexColor("#f0f2f5")

    story = []

    titulo_style = ParagraphStyle("titulo", fontSize=20, textColor=colors.white,
                                   backColor=oscuro, spaceAfter=0, spaceBefore=0,
                                   leftIndent=10, rightIndent=10, leading=28)
    sub_style = ParagraphStyle("sub", fontSize=12, textColor=colors.white,
                                backColor=verde, spaceAfter=12, spaceBefore=0,
                                leftIndent=10, rightIndent=10, leading=20)

    story.append(Paragraph("KYRI — Reporte Diario", titulo_style))
    story.append(Paragraph(business + "  |  " + hoy, sub_style))
    story.append(Spacer(1, 0.4*cm))

    kpi_data = [
        ["INDICADOR", "VALOR"],
        ["Ingresos Totales", "$" + "{:,.0f}".format(total_ingresos) + " COP"],
        ["Ganancia Neta", "$" + "{:,.0f}".format(total_ganancia) + " COP"],
        ["Costo de Ventas", "$" + "{:,.0f}".format(total_costo) + " COP"],
        ["Margen de Ganancia", str(margen) + "%"],
        ["Meta del Día", "$" + "{:,.0f}".format(meta) + " COP"],
        ["Cumplimiento Meta", str(pct_meta) + "%"],
        ["Total Transacciones", str(len(historial))],
    ]

    t = Table(kpi_data, colWidths=[9*cm, 8*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), oscuro),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, gris]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e4e8ef")),
        ("LEFTPADDING", (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 12),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("FONTNAME", (0,1), (0,-1), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph("Productos del Día", ParagraphStyle("h2", fontSize=13,
                textColor=colors.white, backColor=verde, leftIndent=10, leading=22,
                spaceBefore=6, spaceAfter=0)))

    lista = []
    for pid in productos:
        p = productos[pid]
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        vendidos = p.get("sold", 0)
        mg = round((precio - costo) / precio * 100, 1) if precio > 0 else 0
        lista.append([p.get("name", pid), "$" + "{:,.0f}".format(precio),
                      str(vendidos) + " uds", str(mg) + "%", str(p.get("stock", 0)) + " uds"])

    lista.sort(key=lambda x: x[2], reverse=True)
    prod_data = [["Producto", "Precio", "Vendidos", "Margen", "Stock"]] + lista

    t2 = Table(prod_data, colWidths=[6*cm, 3.5*cm, 3*cm, 2.5*cm, 2*cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), oscuro),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, gris]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e4e8ef")),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Generado por KYRI.IA — " + datetime.now().strftime("%d/%m/%Y %I:%M %p"),
                 ParagraphStyle("footer", fontSize=8, textColor=colors.HexColor("#9aa5b4"), alignment=1)))

    doc.build(story)
    buffer.seek(0)
    nombre_archivo = "KYRI_Reporte_" + datetime.now().strftime("%Y-%m-%d") + ".pdf"
    return send_file(buffer, mimetype="application/pdf",
                     as_attachment=True, download_name=nombre_archivo)


# ---------- HISTORIAL ----------

@app.route("/historial")
def historial():
    if not is_logged_in():
        return redirect("/login")

    import glob
    config = load_json("config.json")
    historial_dir = os.path.join(TENANT_PATH, "historial")
    meta_diaria = config.get("daily_goal", 0)

    dias = []
    if os.path.exists(historial_dir):
        archivos = sorted(glob.glob(os.path.join(historial_dir, "*.json")), reverse=True)
        for archivo in archivos:
            with open(archivo, "r", encoding="utf-8") as f:
                data = json.load(f)

            productos = data.get("products", {})
            total_ingresos = 0
            total_ganancia = 0
            total_costo = 0
            for pid in productos:
                p = productos[pid]
                vendidos = p.get("sold", 0)
                precio = p.get("price", 0)
                costo = p.get("cost", 0)
                total_ingresos += vendidos * precio
                total_costo += vendidos * costo
                total_ganancia += vendidos * (precio - costo)

            margen = round(total_ganancia / total_ingresos * 100, 1) if total_ingresos > 0 else 0
            meta = data.get("daily_goal", meta_diaria)
            pct_meta = round(total_ingresos / meta * 100, 1) if meta > 0 else 0
            transacciones = len(data.get("historial_ventas", []))

            fecha_raw = data.get("fecha", "")
            hora_guardado = data.get("hora_guardado", "")
            try:
                from datetime import datetime as dt
                fecha_fmt = dt.strptime(fecha_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                fecha_fmt = fecha_raw

            # nombre del archivo sin extensión para el link de detalle
            nombre_archivo = os.path.basename(archivo).replace(".json", "")

            dias.append({
                "archivo": nombre_archivo,
                "fecha": fecha_fmt,
                "hora_guardado": hora_guardado,
                "fecha_raw": fecha_raw,
                "ingresos": total_ingresos,
                "ganancia": total_ganancia,
                "costo": total_costo,
                "margen": margen,
                "pct_meta": pct_meta,
                "transacciones": transacciones,
                "productos": productos,
                "vs_anterior": 0
            })

    for i in range(len(dias)):
        if i < len(dias) - 1:
            anterior = dias[i + 1]["ingresos"]
            actual = dias[i]["ingresos"]
            if anterior > 0:
                diff = round((actual - anterior) / anterior * 100, 1)
            else:
                diff = 0
            dias[i]["vs_anterior"] = diff

    if dias:
        mejor_dia = max(dias, key=lambda x: x["ingresos"])
        peor_dia = min(dias, key=lambda x: x["ingresos"])
        promedio_ingresos = round(sum(d["ingresos"] for d in dias) / len(dias))
        ganancia_total = sum(d["ganancia"] for d in dias)
        dias_meta_cumplida = len([d for d in dias if d["pct_meta"] >= 100])
        if ganancia_total > 0 and sum(d["ingresos"] for d in dias) > 0:
            margen_promedio = round(ganancia_total / sum(d["ingresos"] for d in dias) * 100, 1)
        else:
            margen_promedio = 0

        if len(dias) >= 14:
            esta_semana = sum(d["ingresos"] for d in dias[:7])
            semana_anterior = sum(d["ingresos"] for d in dias[7:14])
            tendencia = round((esta_semana - semana_anterior) / semana_anterior * 100, 1) if semana_anterior > 0 else 0
        elif len(dias) >= 2:
            mitad = len(dias) // 2
            esta = sum(d["ingresos"] for d in dias[:mitad])
            antes = sum(d["ingresos"] for d in dias[mitad:])
            tendencia = round((esta - antes) / antes * 100, 1) if antes > 0 else 0
        else:
            tendencia = 0

        producto_acum = {}
        for d in dias:
            for pid in d["productos"]:
                p = d["productos"][pid]
                nombre = p.get("name", pid)
                vendidos = p.get("sold", 0)
                producto_acum[nombre] = producto_acum.get(nombre, 0) + vendidos

        top_lista = sorted(producto_acum.items(), key=lambda x: x[1], reverse=True)[:8]
        max_vendidos = top_lista[0][1] if top_lista else 1
        top_productos = [{"nombre": n, "total_vendidos": v, "pct": round(v / max_vendidos * 100)} for n, v in top_lista]
        producto_estrella = top_lista[0][0] if top_lista else "Sin datos"

        dias_grafica = list(reversed(dias[:14]))
        chart_data = {
            "fechas": [d["fecha"] for d in dias_grafica],
            "ingresos": [d["ingresos"] for d in dias_grafica],
            "ganancia": [d["ganancia"] for d in dias_grafica],
            "meta": [d["pct_meta"] for d in dias_grafica]
        }
    else:
        mejor_dia = {"fecha": "-", "ingresos": 0, "margen": 0, "transacciones": 0}
        peor_dia = {"fecha": "-", "ingresos": 0, "margen": 0, "transacciones": 0}
        promedio_ingresos = 0
        ganancia_total = 0
        dias_meta_cumplida = 0
        margen_promedio = 0
        tendencia = 0
        top_productos = []
        producto_estrella = "Sin datos"
        chart_data = {"fechas": [], "ingresos": [], "ganancia": [], "meta": []}

    guardado = request.args.get("guardado", False)

    return render_template("historial.html",
        dias=dias,
        mejor_dia=mejor_dia,
        peor_dia=peor_dia,
        promedio_ingresos=promedio_ingresos,
        ganancia_total=ganancia_total,
        dias_meta_cumplida=dias_meta_cumplida,
        margen_promedio=margen_promedio,
        tendencia=tendencia,
        top_productos=top_productos,
        producto_estrella=producto_estrella,
        chart_data=chart_data,
        business_name=config.get("name", TENANT_ID),
        fecha=datetime.now().strftime("%d/%m/%Y - %I:%M %p"),
        guardado=guardado
    )


@app.route("/historial/guardar_ahora", methods=["POST"])
def guardar_historial_ahora():
    if not is_logged_in():
        return redirect("/login")
    guardar_historial_diario()
    return redirect("/historial?guardado=1")


# ✅ NUEVO: ver detalle de un snapshot específico
@app.route("/historial/dia/<archivo>")
def historial_dia(archivo):
    if not is_logged_in():
        return redirect("/login")
    config = load_json("config.json")
    historial_dir = os.path.join(TENANT_PATH, "historial")
    path = os.path.join(historial_dir, archivo + ".json")
    if not os.path.exists(path):
        return redirect("/historial")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    productos = data.get("products", {})
    total_ingresos = 0
    total_ganancia = 0
    lista_productos = []

    for pid in productos:
        p = productos[pid]
        vendidos = p.get("sold", 0)
        precio = p.get("price", 0)
        costo = p.get("cost", 0)
        ingreso = vendidos * precio
        ganancia = vendidos * (precio - costo)
        total_ingresos += ingreso
        total_ganancia += ganancia
        if vendidos > 0:
            lista_productos.append({
                "nombre": p.get("name", pid),
                "vendidos": vendidos,
                "precio": precio,
                "ingreso": ingreso,
                "ganancia": ganancia
            })

    lista_productos.sort(key=lambda x: x["ingreso"], reverse=True)
    margen = round(total_ganancia / total_ingresos * 100, 1) if total_ingresos > 0 else 0
    meta = data.get("daily_goal", 0)
    pct_meta = round(total_ingresos / meta * 100, 1) if meta > 0 else 0

    return render_template("historial_dia.html",
        fecha=data.get("fecha", ""),
        hora_guardado=data.get("hora_guardado", ""),
        archivo=archivo,
        productos=lista_productos,
        historial_ventas=data.get("historial_ventas", []),
        total_ingresos=total_ingresos,
        total_ganancia=total_ganancia,
        margen=margen,
        meta=meta,
        pct_meta=pct_meta,
        transacciones=len(data.get("historial_ventas", [])),
        business_name=config.get("name", TENANT_ID)
    )

# ---------- RESET DIARIO ----------

import threading
import time

def guardar_historial_diario():
    hoy = datetime.now().strftime("%Y-%m-%d")
    historial_dir = os.path.join(TENANT_PATH, "historial")
    if not os.path.exists(historial_dir):
        os.makedirs(historial_dir)
    state = load_json("state.json")
    config = load_json("config.json")
    snapshot = {
        "fecha": hoy,
        "hora_guardado": datetime.now().strftime("%I:%M %p"),
        "business_name": config.get("name", TENANT_ID),
        "products": state.get("products", {}),
        "sales": state.get("sales", 0),
        "historial_ventas": state.get("historial_ventas", []),
        "daily_goal": config.get("daily_goal", 0)
    }
    path = os.path.join(historial_dir, hoy + ".json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)
    print("Historial guardado: " + path)

def reset_diario():
    while True:
        ahora = datetime.now()
        from datetime import timedelta
        hoy_7pm = ahora.replace(hour=19, minute=0, second=0, microsecond=0)
        if ahora >= hoy_7pm:
            proxima = hoy_7pm + timedelta(days=1)
        else:
            proxima = hoy_7pm
        segundos = (proxima - ahora).total_seconds()
        time.sleep(segundos)
        guardar_historial_diario()
        state = load_json("state.json")
        productos = state.get("products", {})
        for pid in productos:
            productos[pid]["sold"] = 0
        state["products"] = productos
        state["sales"] = 0
        state["historial_ventas"] = []
        state["insights"] = []
        save_json("state.json", state)
        print("Reset 7PM ejecutado: " + datetime.now().strftime("%d/%m/%Y %H:%M"))

hilo_reset = threading.Thread(target=reset_diario, daemon=True)
hilo_reset.start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
